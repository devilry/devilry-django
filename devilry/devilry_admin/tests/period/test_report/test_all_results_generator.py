# -*- coding: utf-8 -*-

import mock
import openpyxl
from django import test
from django.conf import settings
from django.utils import timezone

from model_bakery import baker
from io import BytesIO

from devilry.devilry_dbcache.customsql import AssignmentGroupDbCacheCustomSql
from devilry.devilry_group import devilry_group_baker_factories as group_factory
from devilry.devilry_admin.views.period import all_results_generator
from devilry.devilry_report.models import DevilryReport


class AllResultsGeneratorPreMixin:
    def get_work_sheet(self, devilry_report):
        raise NotImplementedError()

    def make_assignment(self, period, **assignment_kwargs):
        return baker.make('core.Assignment', parentnode=period, **assignment_kwargs)

    def make_relatedstudent(self, period, **relatedstudent_kwargs):
        return baker.make('core.RelatedStudent', period=period, **relatedstudent_kwargs)

    def make_group_for_student(self, assignment, relatedstudent):
        group = baker.make('core.AssignmentGroup', parentnode=assignment)
        baker.make('core.Candidate', assignment_group=group, relatedstudent=relatedstudent)
        return group

    def test_headings(self):
        requestuser = baker.make(settings.AUTH_USER_MODEL)
        period = baker.make_recipe('devilry.apps.core.period_active')
        testassignment1 = self.make_assignment(period=period, long_name='Assignment 1')
        testassignment2 = self.make_assignment(period=period, long_name='Assignment 2')
        testassignment3 = self.make_assignment(period=period, long_name='Assignment 3')
        testassignment4 = self.make_assignment(period=period, long_name='Assignment 4')

        with mock.patch.object(DevilryReport, 'generator', all_results_generator.AllResultsExcelReportGenerator):
            devilry_report = DevilryReport(
                generator_options={'period_id': period.id},
                generator_type='semesterstudentresults',
                generated_by_user=requestuser)
            devilry_report.full_clean()
            devilry_report.save()
            devilry_report.generate()
            worksheet = self.get_work_sheet(devilry_report=devilry_report)
            self.assertEqual(worksheet.cell(row=1, column=1).value, 'Student')
            self.assertEqual(worksheet.cell(row=1, column=2).value, testassignment1.long_name)
            self.assertEqual(worksheet.cell(row=1, column=3).value, testassignment2.long_name)
            self.assertEqual(worksheet.cell(row=1, column=4).value, testassignment3.long_name)
            self.assertEqual(worksheet.cell(row=1, column=5).value, testassignment4.long_name)

    def create_single_student_passed_plugin_passed_failed(self):
        period = baker.make_recipe('devilry.apps.core.period_active')
        testassignment = self.make_assignment(
            period=period, long_name='Assignment 1', passing_grade_min_points=1, max_points=1)
        teststudent = self.make_relatedstudent(period=period, user__shortname='teststudent@example.com')
        testgroup = self.make_group_for_student(assignment=testassignment, relatedstudent=teststudent)
        group_factory.feedbackset_first_attempt_published(group=testgroup, grading_points=1)
        return period

    def create_single_student_failed_plugin_passed_failed(self):
        period = baker.make_recipe('devilry.apps.core.period_active')
        testassignment = self.make_assignment(
            period=period, long_name='Assignment 1', passing_grade_min_points=1, max_points=1)
        teststudent = self.make_relatedstudent(period=period, user__shortname='teststudent@example.com')
        testgroup = self.make_group_for_student(assignment=testassignment, relatedstudent=teststudent)
        group_factory.feedbackset_first_attempt_published(group=testgroup, grading_points=0)
        return period

    def create_single_student_plugin_raw_points(self):
        period = baker.make_recipe('devilry.apps.core.period_active')
        testassignment = self.make_assignment(
            period=period, long_name='Assignment 1', passing_grade_min_points=5, max_points=10,
            points_to_grade_mapper='raw-points')
        teststudent = self.make_relatedstudent(period=period, user__shortname='teststudent@example.com')
        testgroup = self.make_group_for_student(assignment=testassignment, relatedstudent=teststudent)
        group_factory.feedbackset_first_attempt_published(group=testgroup, grading_points=7)
        return period

    def create_single_student_plugin_custom_table(self):
        period = baker.make_recipe('devilry.apps.core.period_active')
        testassignment = self.make_assignment(
            period=period, long_name='Assignment 1', passing_grade_min_points=5, max_points=10,
            points_to_grade_mapper='custom-table')
        point_to_grade_map = baker.make('core.PointToGradeMap', assignment=testassignment)
        baker.make('core.PointRangeToGrade',point_to_grade_map=point_to_grade_map,
                   minimum_points=0, maximum_points=10, grade='F')
        baker.make('core.PointRangeToGrade', point_to_grade_map=point_to_grade_map,
                   minimum_points=11, maximum_points=20, grade='E')
        baker.make('core.PointRangeToGrade', point_to_grade_map=point_to_grade_map,
                   minimum_points=21, maximum_points=30, grade='D')
        baker.make('core.PointRangeToGrade', point_to_grade_map=point_to_grade_map,
                   minimum_points=31, maximum_points=40, grade='C')
        baker.make('core.PointRangeToGrade', point_to_grade_map=point_to_grade_map,
                   minimum_points=41, maximum_points=50, grade='B')
        baker.make('core.PointRangeToGrade', point_to_grade_map=point_to_grade_map,
                   minimum_points=51, maximum_points=60, grade='A')
        teststudent = self.make_relatedstudent(period=period, user__shortname='teststudent@example.com')
        testgroup = self.make_group_for_student(assignment=testassignment, relatedstudent=teststudent)
        group_factory.feedbackset_first_attempt_published(group=testgroup, grading_points=32)
        return period

    def create_single_student_not_registered_on_assignment(self):
        period = baker.make_recipe('devilry.apps.core.period_active')
        self.make_assignment(period=period, long_name='Assignment 1', passing_grade_min_points=1, max_points=1)
        self.make_relatedstudent(period=period, user__shortname='teststudent@example.com')
        return period

    def create_single_student_waiting_for_feedback(self):
        period = baker.make_recipe('devilry.apps.core.period_active')
        testassignment = self.make_assignment(period=period, long_name='Assignment 1', passing_grade_min_points=1, max_points=1)
        teststudent = self.make_relatedstudent(period=period, user__shortname='teststudent@example.com')
        testgroup = self.make_group_for_student(assignment=testassignment, relatedstudent=teststudent)
        deadline_datetime = timezone.now() - timezone.timedelta(days=10)
        group_factory.feedbackset_first_attempt_unpublished(group=testgroup, deadline_datetime=deadline_datetime)
        return period

    def create_single_student_hard_deadline_no_deliveries(self):
        period = baker.make_recipe('devilry.apps.core.period_active')
        testassignment = self.make_assignment(
            period=period, long_name='Assignment 1', passing_grade_min_points=1, max_points=1, deadline_handling=1)
        teststudent = self.make_relatedstudent(period=period, user__shortname='teststudent@example.com')
        testgroup = self.make_group_for_student(assignment=testassignment, relatedstudent=teststudent)
        deadline_datetime = timezone.now() - timezone.timedelta(days=10)
        group_factory.feedbackset_first_attempt_unpublished(group=testgroup, deadline_datetime=deadline_datetime)
        return period

    def create_single_student_hard_deadline_waiting_for_feedback(self):
        period = baker.make_recipe('devilry.apps.core.period_active')
        testassignment = self.make_assignment(
            period=period, long_name='Assignment 1', passing_grade_min_points=1, max_points=1, deadline_handling=1)
        teststudent = self.make_relatedstudent(period=period, user__shortname='teststudent@example.com')
        testgroup = self.make_group_for_student(assignment=testassignment, relatedstudent=teststudent)
        deadline_datetime = timezone.now() - timezone.timedelta(days=10)
        feedbackset = group_factory.feedbackset_first_attempt_unpublished(group=testgroup, deadline_datetime=deadline_datetime)
        baker.make('devilry_group.GroupComment', feedback_set=feedbackset, user=teststudent.user,
                   user_role='student', text='Test')
        return period

    def create_single_student_waiting_for_deliveries(self):
        period = baker.make_recipe('devilry.apps.core.period_active')
        testassignment = self.make_assignment(
            period=period, long_name='Assignment 1', passing_grade_min_points=1, max_points=1, deadline_handling=1)
        teststudent = self.make_relatedstudent(period=period, user__shortname='teststudent@example.com')
        testgroup = self.make_group_for_student(assignment=testassignment, relatedstudent=teststudent)
        deadline_datetime = timezone.now() + timezone.timedelta(days=10)
        group_factory.feedbackset_first_attempt_unpublished(group=testgroup, deadline_datetime=deadline_datetime)
        return period

    def create_single_student_on_multiple_assignments(self):
        period = baker.make_recipe('devilry.apps.core.period_active')
        testassignment1 = self.make_assignment(
            period=period, long_name='Assignment 1', passing_grade_min_points=1, max_points=1)
        testassignment2 = self.make_assignment(
            period=period, long_name='Assignment 1', passing_grade_min_points=1, max_points=1)
        testassignment3 = self.make_assignment(
            period=period, long_name='Assignment 1', passing_grade_min_points=1, max_points=1)
        teststudent = self.make_relatedstudent(period=period, user__shortname='teststudent@example.com')
        testgroup1 = self.make_group_for_student(assignment=testassignment1, relatedstudent=teststudent)
        testgroup2 = self.make_group_for_student(assignment=testassignment2, relatedstudent=teststudent)
        testgroup3 = self.make_group_for_student(assignment=testassignment3, relatedstudent=teststudent)
        group_factory.feedbackset_first_attempt_published(group=testgroup1, grading_points=1)
        group_factory.feedbackset_first_attempt_published(group=testgroup2, grading_points=1)
        group_factory.feedbackset_first_attempt_published(group=testgroup3, grading_points=0)
        return period

    def create_multiple_students_on_multiple_assignments(self):
        period = baker.make_recipe('devilry.apps.core.period_active')
        testassignment1 = self.make_assignment(
            period=period, long_name='Assignment 1', passing_grade_min_points=1, max_points=1)
        testassignment2 = self.make_assignment(
            period=period, long_name='Assignment 1', passing_grade_min_points=1, max_points=1)
        teststudent1 = self.make_relatedstudent(period=period, user__shortname='teststudent1@example.com')
        teststudent2 = self.make_relatedstudent(period=period, user__shortname='teststudent2@example.com')
        testgroup1 = self.make_group_for_student(assignment=testassignment1, relatedstudent=teststudent1)
        testgroup2 = self.make_group_for_student(assignment=testassignment2, relatedstudent=teststudent1)
        testgroup3 = self.make_group_for_student(assignment=testassignment1, relatedstudent=teststudent2)
        testgroup4 = self.make_group_for_student(assignment=testassignment2, relatedstudent=teststudent2)
        group_factory.feedbackset_first_attempt_published(group=testgroup1, grading_points=1)
        group_factory.feedbackset_first_attempt_published(group=testgroup2, grading_points=1)
        group_factory.feedbackset_first_attempt_published(group=testgroup3, grading_points=1)
        group_factory.feedbackset_first_attempt_published(group=testgroup4, grading_points=0)
        return period

    def create_single_student_results_on_each_points_to_grade_mapper_type_assignment(self):
        period = baker.make_recipe('devilry.apps.core.period_active')
        testassignment_passed_failed = self.make_assignment(
            period=period, long_name='Assignment Passed Failed', passing_grade_min_points=5, max_points=10)
        testassignment_raw_points = self.make_assignment(
            period=period, long_name='Assignment Raw Points', passing_grade_min_points=5, max_points=10,
            points_to_grade_mapper='raw-points')
        testassignment_custom_table = self.make_assignment(
            period=period, long_name='Assignment Raw Points', passing_grade_min_points=10, max_points=60,
            points_to_grade_mapper='custom-table')
        point_to_grade_map = baker.make('core.PointToGradeMap', assignment=testassignment_custom_table)
        baker.make('core.PointRangeToGrade', point_to_grade_map=point_to_grade_map,
                   minimum_points=0, maximum_points=10, grade='F')
        baker.make('core.PointRangeToGrade', point_to_grade_map=point_to_grade_map,
                   minimum_points=11, maximum_points=20, grade='E')
        baker.make('core.PointRangeToGrade', point_to_grade_map=point_to_grade_map,
                   minimum_points=21, maximum_points=30, grade='D')
        baker.make('core.PointRangeToGrade', point_to_grade_map=point_to_grade_map,
                   minimum_points=31, maximum_points=40, grade='C')
        baker.make('core.PointRangeToGrade', point_to_grade_map=point_to_grade_map,
                   minimum_points=41, maximum_points=50, grade='B')
        baker.make('core.PointRangeToGrade', point_to_grade_map=point_to_grade_map,
                   minimum_points=51, maximum_points=60, grade='A')
        teststudent = self.make_relatedstudent(period=period, user__shortname='teststudent@example.com')
        testgroup_passed_failed = self.make_group_for_student(
            assignment=testassignment_passed_failed, relatedstudent=teststudent)
        group_factory.feedbackset_first_attempt_published(group=testgroup_passed_failed, grading_points=5)
        testgroup_raw_points = self.make_group_for_student(
            assignment=testassignment_raw_points, relatedstudent=teststudent)
        group_factory.feedbackset_first_attempt_published(group=testgroup_raw_points, grading_points=5)
        testgroup_custom_table = self.make_group_for_student(
            assignment=testassignment_custom_table, relatedstudent=teststudent)
        group_factory.feedbackset_first_attempt_published(group=testgroup_custom_table, grading_points=47)
        return period


class TestAllResultsGeneratorGradesSheet(AllResultsGeneratorPreMixin, test.TestCase):
    def setUp(self):
        AssignmentGroupDbCacheCustomSql().initialize()

    def get_work_sheet(self, devilry_report):
        return openpyxl.load_workbook(filename=BytesIO(devilry_report.result)).get_sheet_by_name(name='Grades')

    def generate_report_and_get_worksheet(self, generated_by_user, period, worksheet_name):
        devilry_report = DevilryReport(
            generator_options={'period_id': period.id},
            generator_type='semesterstudentresults',
            generated_by_user=generated_by_user)
        devilry_report.full_clean()
        devilry_report.save()
        devilry_report.generate()
        return openpyxl.load_workbook(filename=BytesIO(devilry_report.result))\
            .get_sheet_by_name(name=worksheet_name)

    def test_grades_sheet_value_for_single_student_passed_plugin_passed(self):
        requestuser = baker.make(settings.AUTH_USER_MODEL)
        period = self.create_single_student_passed_plugin_passed_failed()

        with mock.patch.object(DevilryReport, 'generator', all_results_generator.AllResultsExcelReportGenerator):
            worksheet = self.generate_report_and_get_worksheet(
                generated_by_user=requestuser, period=period, worksheet_name='Grades')
            self.assertEqual(worksheet.cell(row=2, column=1).value, 'teststudent@example.com')
            self.assertEqual(worksheet.cell(row=2, column=2).value, 'passed')

    def test_points_sheet_value_for_single_student_passed_plugin_passed(self):
        requestuser = baker.make(settings.AUTH_USER_MODEL)
        period = self.create_single_student_passed_plugin_passed_failed()

        with mock.patch.object(DevilryReport, 'generator', all_results_generator.AllResultsExcelReportGenerator):
            worksheet = self.generate_report_and_get_worksheet(
                generated_by_user=requestuser, period=period, worksheet_name='Points')
            self.assertEqual(worksheet.cell(row=2, column=1).value, 'teststudent@example.com')
            self.assertEqual(worksheet.cell(row=2, column=2).value, 1)

    def test_passed_sheet_value_for_single_student_passed_plugin_passed(self):
        requestuser = baker.make(settings.AUTH_USER_MODEL)
        period = self.create_single_student_passed_plugin_passed_failed()

        with mock.patch.object(DevilryReport, 'generator', all_results_generator.AllResultsExcelReportGenerator):
            worksheet = self.generate_report_and_get_worksheet(
                generated_by_user=requestuser, period=period, worksheet_name='Passed Failed')
            self.assertEqual(worksheet.cell(row=2, column=1).value, 'teststudent@example.com')
            self.assertEqual(worksheet.cell(row=2, column=2).value, 1)

    def test_grades_sheet_value_for_single_student_failed_plugin_passed(self):
        requestuser = baker.make(settings.AUTH_USER_MODEL)
        period = self.create_single_student_failed_plugin_passed_failed()

        with mock.patch.object(DevilryReport, 'generator', all_results_generator.AllResultsExcelReportGenerator):
            worksheet = self.generate_report_and_get_worksheet(
                generated_by_user=requestuser, period=period, worksheet_name='Grades')
            self.assertEqual(worksheet.cell(row=2, column=1).value, 'teststudent@example.com')
            self.assertEqual(worksheet.cell(row=2, column=2).value, 'failed')

    def test_points_sheet_value_for_single_student_failed_plugin_passed(self):
        requestuser = baker.make(settings.AUTH_USER_MODEL)
        period = self.create_single_student_failed_plugin_passed_failed()

        with mock.patch.object(DevilryReport, 'generator', all_results_generator.AllResultsExcelReportGenerator):
            worksheet = self.generate_report_and_get_worksheet(
                generated_by_user=requestuser, period=period, worksheet_name='Points')
            self.assertEqual(worksheet.cell(row=2, column=1).value, 'teststudent@example.com')
            self.assertEqual(worksheet.cell(row=2, column=2).value, 0)

    def test_passed_sheet_value_for_single_student_failed_plugin_passed(self):
        requestuser = baker.make(settings.AUTH_USER_MODEL)
        period = self.create_single_student_failed_plugin_passed_failed()

        with mock.patch.object(DevilryReport, 'generator', all_results_generator.AllResultsExcelReportGenerator):
            worksheet = self.generate_report_and_get_worksheet(
                generated_by_user=requestuser, period=period, worksheet_name='Passed Failed')
            self.assertEqual(worksheet.cell(row=2, column=1).value, 'teststudent@example.com')
            self.assertEqual(worksheet.cell(row=2, column=2).value, 0)

    def test_grades_sheet_value_for_single_student_plugin_raw_points(self):
        requestuser = baker.make(settings.AUTH_USER_MODEL)
        period = self.create_single_student_plugin_raw_points()

        with mock.patch.object(DevilryReport, 'generator', all_results_generator.AllResultsExcelReportGenerator):
            worksheet = self.generate_report_and_get_worksheet(
                generated_by_user=requestuser, period=period, worksheet_name='Grades')
            self.assertEqual(worksheet.cell(row=2, column=1).value, 'teststudent@example.com')
            self.assertEqual(worksheet.cell(row=2, column=2).value, '7/10')

    def test_points_sheet_value_for_single_student_plugin_raw_points(self):
        requestuser = baker.make(settings.AUTH_USER_MODEL)
        period = self.create_single_student_plugin_raw_points()

        with mock.patch.object(DevilryReport, 'generator', all_results_generator.AllResultsExcelReportGenerator):
            worksheet = self.generate_report_and_get_worksheet(
                generated_by_user=requestuser, period=period, worksheet_name='Points')
            self.assertEqual(worksheet.cell(row=2, column=1).value, 'teststudent@example.com')
            self.assertEqual(worksheet.cell(row=2, column=2).value, 7)

    def test_passed_sheet_value_for_single_student_plugin_raw_points(self):
        requestuser = baker.make(settings.AUTH_USER_MODEL)
        period = self.create_single_student_plugin_raw_points()

        with mock.patch.object(DevilryReport, 'generator', all_results_generator.AllResultsExcelReportGenerator):
            worksheet = self.generate_report_and_get_worksheet(
                generated_by_user=requestuser, period=period, worksheet_name='Passed Failed')
            self.assertEqual(worksheet.cell(row=2, column=1).value, 'teststudent@example.com')
            self.assertEqual(worksheet.cell(row=2, column=2).value, 1)

    def test_grades_sheet_value_for_single_student_plugin_custom_table(self):
        requestuser = baker.make(settings.AUTH_USER_MODEL)
        period = self.create_single_student_plugin_custom_table()

        with mock.patch.object(DevilryReport, 'generator', all_results_generator.AllResultsExcelReportGenerator):
            worksheet = self.generate_report_and_get_worksheet(
                generated_by_user=requestuser, period=period, worksheet_name='Grades')
            self.assertEqual(worksheet.cell(row=2, column=1).value, 'teststudent@example.com')
            self.assertEqual(worksheet.cell(row=2, column=2).value, 'C')

    def test_points_sheet_value_for_single_student_plugin_custom_table(self):
        requestuser = baker.make(settings.AUTH_USER_MODEL)
        period = self.create_single_student_plugin_custom_table()

        with mock.patch.object(DevilryReport, 'generator', all_results_generator.AllResultsExcelReportGenerator):
            worksheet = self.generate_report_and_get_worksheet(
                generated_by_user=requestuser, period=period, worksheet_name='Points')
            self.assertEqual(worksheet.cell(row=2, column=1).value, 'teststudent@example.com')
            self.assertEqual(worksheet.cell(row=2, column=2).value, 32)

    def test_passed_sheet_value_for_single_student_plugin_custom_table(self):
        requestuser = baker.make(settings.AUTH_USER_MODEL)
        period = self.create_single_student_plugin_custom_table()

        with mock.patch.object(DevilryReport, 'generator', all_results_generator.AllResultsExcelReportGenerator):
            worksheet = self.generate_report_and_get_worksheet(
                generated_by_user=requestuser, period=period, worksheet_name='Passed Failed')
            self.assertEqual(worksheet.cell(row=2, column=1).value, 'teststudent@example.com')
            self.assertEqual(worksheet.cell(row=2, column=2).value, 1)

    def test_grades_sheet_value_for_single_student_not_registered_on_assignment(self):
        requestuser = baker.make(settings.AUTH_USER_MODEL)
        period = self.create_single_student_not_registered_on_assignment()

        with mock.patch.object(DevilryReport, 'generator', all_results_generator.AllResultsExcelReportGenerator):
            worksheet = self.generate_report_and_get_worksheet(
                generated_by_user=requestuser, period=period, worksheet_name='Grades')
            self.assertEqual(worksheet.cell(row=2, column=1).value, 'teststudent@example.com')
            self.assertEqual(worksheet.cell(row=2, column=2).value, 'Not registered')

    def test_points_sheet_value_for_single_student_not_registered_on_assignment(self):
        requestuser = baker.make(settings.AUTH_USER_MODEL)
        period = self.create_single_student_not_registered_on_assignment()

        with mock.patch.object(DevilryReport, 'generator', all_results_generator.AllResultsExcelReportGenerator):
            worksheet = self.generate_report_and_get_worksheet(
                generated_by_user=requestuser, period=period, worksheet_name='Points')
            self.assertEqual(worksheet.cell(row=2, column=1).value, 'teststudent@example.com')
            self.assertEqual(worksheet.cell(row=2, column=2).value, None)

    def test_passed_sheet_value_for_single_student_not_registered_on_assignment(self):
        requestuser = baker.make(settings.AUTH_USER_MODEL)
        period = self.create_single_student_not_registered_on_assignment()

        with mock.patch.object(DevilryReport, 'generator', all_results_generator.AllResultsExcelReportGenerator):
            worksheet = self.generate_report_and_get_worksheet(
                generated_by_user=requestuser, period=period, worksheet_name='Passed Failed')
            self.assertEqual(worksheet.cell(row=2, column=1).value, 'teststudent@example.com')
            self.assertEqual(worksheet.cell(row=2, column=2).value, None)

    def test_grades_sheet_value_for_single_student_waiting_for_feedback(self):
        requestuser = baker.make(settings.AUTH_USER_MODEL)
        period = self.create_single_student_waiting_for_feedback()

        with mock.patch.object(DevilryReport, 'generator', all_results_generator.AllResultsExcelReportGenerator):
            worksheet = self.generate_report_and_get_worksheet(
                generated_by_user=requestuser, period=period, worksheet_name='Grades')
            self.assertEqual(worksheet.cell(row=2, column=1).value, 'teststudent@example.com')
            self.assertEqual(worksheet.cell(row=2, column=2).value, 'Waiting for feedback')

    def test_points_sheet_value_for_single_student_waiting_for_feedback(self):
        requestuser = baker.make(settings.AUTH_USER_MODEL)
        period = self.create_single_student_waiting_for_feedback()

        with mock.patch.object(DevilryReport, 'generator', all_results_generator.AllResultsExcelReportGenerator):
            worksheet = self.generate_report_and_get_worksheet(
                generated_by_user=requestuser, period=period, worksheet_name='Points')
            self.assertEqual(worksheet.cell(row=2, column=1).value, 'teststudent@example.com')
            self.assertEqual(worksheet.cell(row=2, column=2).value, None)

    def test_passed_sheet_value_for_single_student_waiting_for_feedback(self):
        requestuser = baker.make(settings.AUTH_USER_MODEL)
        period = self.create_single_student_waiting_for_feedback()

        with mock.patch.object(DevilryReport, 'generator', all_results_generator.AllResultsExcelReportGenerator):
            worksheet = self.generate_report_and_get_worksheet(
                generated_by_user=requestuser, period=period, worksheet_name='Passed Failed')
            self.assertEqual(worksheet.cell(row=2, column=1).value, 'teststudent@example.com')
            self.assertEqual(worksheet.cell(row=2, column=2).value, None)

    def test_grades_sheet_value_for_single_student_hard_deadline_no_deliveries(self):
        requestuser = baker.make(settings.AUTH_USER_MODEL)
        period = self.create_single_student_hard_deadline_no_deliveries()

        with mock.patch.object(DevilryReport, 'generator', all_results_generator.AllResultsExcelReportGenerator):
            worksheet = self.generate_report_and_get_worksheet(
                generated_by_user=requestuser, period=period, worksheet_name='Grades')
            self.assertEqual(worksheet.cell(row=2, column=1).value, 'teststudent@example.com')
            self.assertEqual(worksheet.cell(row=2, column=2).value, 'No deliveries')

    def test_points_sheet_value_for_single_student_hard_deadline_no_deliveries(self):
        requestuser = baker.make(settings.AUTH_USER_MODEL)
        period = self.create_single_student_hard_deadline_no_deliveries()

        with mock.patch.object(DevilryReport, 'generator', all_results_generator.AllResultsExcelReportGenerator):
            worksheet = self.generate_report_and_get_worksheet(
                generated_by_user=requestuser, period=period, worksheet_name='Points')
            self.assertEqual(worksheet.cell(row=2, column=1).value, 'teststudent@example.com')
            self.assertEqual(worksheet.cell(row=2, column=2).value, None)

    def test_passed_sheet_value_for_single_student_hard_deadline_no_deliveries(self):
        requestuser = baker.make(settings.AUTH_USER_MODEL)
        period = self.create_single_student_hard_deadline_no_deliveries()

        with mock.patch.object(DevilryReport, 'generator', all_results_generator.AllResultsExcelReportGenerator):
            worksheet = self.generate_report_and_get_worksheet(
                generated_by_user=requestuser, period=period, worksheet_name='Passed Failed')
            self.assertEqual(worksheet.cell(row=2, column=1).value, 'teststudent@example.com')
            self.assertEqual(worksheet.cell(row=2, column=2).value, None)

    def test_grades_sheet_value_for_single_student_hard_deadline_waiting_for_feedback(self):
        requestuser = baker.make(settings.AUTH_USER_MODEL)
        period = self.create_single_student_hard_deadline_waiting_for_feedback()

        with mock.patch.object(DevilryReport, 'generator', all_results_generator.AllResultsExcelReportGenerator):
            worksheet = self.generate_report_and_get_worksheet(
                generated_by_user=requestuser, period=period, worksheet_name='Grades')
            self.assertEqual(worksheet.cell(row=2, column=1).value, 'teststudent@example.com')
            self.assertEqual(worksheet.cell(row=2, column=2).value, 'Waiting for feedback')

    def test_points_sheet_value_for_single_student_hard_deadline_waiting_for_feedback(self):
        requestuser = baker.make(settings.AUTH_USER_MODEL)
        period = self.create_single_student_hard_deadline_waiting_for_feedback()

        with mock.patch.object(DevilryReport, 'generator', all_results_generator.AllResultsExcelReportGenerator):
            worksheet = self.generate_report_and_get_worksheet(
                generated_by_user=requestuser, period=period, worksheet_name='Points')
            self.assertEqual(worksheet.cell(row=2, column=1).value, 'teststudent@example.com')
            self.assertEqual(worksheet.cell(row=2, column=2).value, None)

    def test_passed_sheet_value_for_single_student_hard_deadline_waiting_for_feedback(self):
        requestuser = baker.make(settings.AUTH_USER_MODEL)
        period = self.create_single_student_hard_deadline_waiting_for_feedback()

        with mock.patch.object(DevilryReport, 'generator', all_results_generator.AllResultsExcelReportGenerator):
            worksheet = self.generate_report_and_get_worksheet(
                generated_by_user=requestuser, period=period, worksheet_name='Passed Failed')
            self.assertEqual(worksheet.cell(row=2, column=1).value, 'teststudent@example.com')
            self.assertEqual(worksheet.cell(row=2, column=2).value, None)

    def test_grades_sheet_value_for_single_student_waiting_for_deliveries(self):
        requestuser = baker.make(settings.AUTH_USER_MODEL)
        period = self.create_single_student_waiting_for_deliveries()

        with mock.patch.object(DevilryReport, 'generator', all_results_generator.AllResultsExcelReportGenerator):
            worksheet = self.generate_report_and_get_worksheet(
                generated_by_user=requestuser, period=period, worksheet_name='Grades')
            self.assertEqual(worksheet.cell(row=2, column=1).value, 'teststudent@example.com')
            self.assertEqual(worksheet.cell(row=2, column=2).value, 'Waiting for deliveries')

    def test_points_sheet_value_for_single_student_waiting_for_deliveries(self):
        requestuser = baker.make(settings.AUTH_USER_MODEL)
        period = self.create_single_student_waiting_for_deliveries()

        with mock.patch.object(DevilryReport, 'generator', all_results_generator.AllResultsExcelReportGenerator):
            worksheet = self.generate_report_and_get_worksheet(
                generated_by_user=requestuser, period=period, worksheet_name='Points')
            self.assertEqual(worksheet.cell(row=2, column=1).value, 'teststudent@example.com')
            self.assertEqual(worksheet.cell(row=2, column=2).value, None)

    def test_passed_sheet_value_for_single_student_waiting_for_deliveries(self):
        requestuser = baker.make(settings.AUTH_USER_MODEL)
        period = self.create_single_student_waiting_for_deliveries()

        with mock.patch.object(DevilryReport, 'generator', all_results_generator.AllResultsExcelReportGenerator):
            worksheet = self.generate_report_and_get_worksheet(
                generated_by_user=requestuser, period=period, worksheet_name='Passed Failed')
            self.assertEqual(worksheet.cell(row=2, column=1).value, 'teststudent@example.com')
            self.assertEqual(worksheet.cell(row=2, column=2).value, None)

    def test_grades_sheet_value_for_single_student_on_multiple_assignments(self):
        requestuser = baker.make(settings.AUTH_USER_MODEL)
        period = self.create_single_student_on_multiple_assignments()

        with mock.patch.object(DevilryReport, 'generator', all_results_generator.AllResultsExcelReportGenerator):
            worksheet = self.generate_report_and_get_worksheet(
                generated_by_user=requestuser, period=period, worksheet_name='Grades')
            self.assertEqual(worksheet.cell(row=2, column=1).value, 'teststudent@example.com')
            self.assertEqual(worksheet.cell(row=2, column=2).value, 'passed')
            self.assertEqual(worksheet.cell(row=2, column=3).value, 'passed')
            self.assertEqual(worksheet.cell(row=2, column=4).value, 'failed')

    def test_points_sheet_value_for_single_student_on_multiple_assignments(self):
        requestuser = baker.make(settings.AUTH_USER_MODEL)
        period = self.create_single_student_on_multiple_assignments()

        with mock.patch.object(DevilryReport, 'generator', all_results_generator.AllResultsExcelReportGenerator):
            worksheet = self.generate_report_and_get_worksheet(
                generated_by_user=requestuser, period=period, worksheet_name='Points')
            self.assertEqual(worksheet.cell(row=2, column=1).value, 'teststudent@example.com')
            self.assertEqual(worksheet.cell(row=2, column=2).value, 1)
            self.assertEqual(worksheet.cell(row=2, column=3).value, 1)
            self.assertEqual(worksheet.cell(row=2, column=4).value, 0)

    def test_passed_sheet_value_for_single_student_on_multiple_assignments(self):
        requestuser = baker.make(settings.AUTH_USER_MODEL)
        period = self.create_single_student_on_multiple_assignments()

        with mock.patch.object(DevilryReport, 'generator', all_results_generator.AllResultsExcelReportGenerator):
            worksheet = self.generate_report_and_get_worksheet(
                generated_by_user=requestuser, period=period, worksheet_name='Passed Failed')
            self.assertEqual(worksheet.cell(row=2, column=1).value, 'teststudent@example.com')
            self.assertEqual(worksheet.cell(row=2, column=2).value, 1)
            self.assertEqual(worksheet.cell(row=2, column=3).value, 1)
            self.assertEqual(worksheet.cell(row=2, column=4).value, 0)

    def test_grades_sheet_multiple_students_on_multiple_assignments(self):
        requestuser = baker.make(settings.AUTH_USER_MODEL)
        period = self.create_multiple_students_on_multiple_assignments()

        with mock.patch.object(DevilryReport, 'generator', all_results_generator.AllResultsExcelReportGenerator):
            worksheet = self.generate_report_and_get_worksheet(
                generated_by_user=requestuser, period=period, worksheet_name='Grades')
            if worksheet.cell(row=2, column=1).value == 'teststudent1@example.com':
                teststudent1_row, teststudent2_row = (2, 3)
            else:
                teststudent1_row, teststudent2_row = (3, 2)
            self.assertEqual(worksheet.cell(row=teststudent1_row, column=1).value, 'teststudent1@example.com')
            self.assertEqual(worksheet.cell(row=teststudent1_row, column=2).value, 'passed')
            self.assertEqual(worksheet.cell(row=teststudent1_row, column=3).value, 'passed')
            self.assertEqual(worksheet.cell(row=teststudent2_row, column=1).value, 'teststudent2@example.com')
            self.assertEqual(worksheet.cell(row=teststudent2_row, column=2).value, 'passed')
            self.assertEqual(worksheet.cell(row=teststudent2_row, column=3).value, 'failed')

    def test_points_sheet_multiple_students_on_multiple_assignments(self):
        requestuser = baker.make(settings.AUTH_USER_MODEL)
        period = self.create_multiple_students_on_multiple_assignments()

        with mock.patch.object(DevilryReport, 'generator', all_results_generator.AllResultsExcelReportGenerator):
            worksheet = self.generate_report_and_get_worksheet(
                generated_by_user=requestuser, period=period, worksheet_name='Points')
            if worksheet.cell(row=2, column=1).value == 'teststudent1@example.com':
                teststudent1_row, teststudent2_row = (2, 3)
            else:
                teststudent1_row, teststudent2_row = (3, 2)
            self.assertEqual(worksheet.cell(row=teststudent1_row, column=1).value, 'teststudent1@example.com')
            self.assertEqual(worksheet.cell(row=teststudent1_row, column=2).value, 1)
            self.assertEqual(worksheet.cell(row=teststudent1_row, column=3).value, 1)
            self.assertEqual(worksheet.cell(row=teststudent2_row, column=1).value, 'teststudent2@example.com')
            self.assertEqual(worksheet.cell(row=teststudent2_row, column=2).value, 1)
            self.assertEqual(worksheet.cell(row=teststudent2_row, column=3).value, 0)

    def test_passed_sheet_multiple_students_on_multiple_assignments(self):
        requestuser = baker.make(settings.AUTH_USER_MODEL)
        period = self.create_multiple_students_on_multiple_assignments()

        with mock.patch.object(DevilryReport, 'generator', all_results_generator.AllResultsExcelReportGenerator):
            worksheet = self.generate_report_and_get_worksheet(
                generated_by_user=requestuser, period=period, worksheet_name='Passed Failed')
            if worksheet.cell(row=2, column=1).value == 'teststudent1@example.com':
                teststudent1_row, teststudent2_row = (2, 3)
            else:
                teststudent1_row, teststudent2_row = (3, 2)
            self.assertEqual(worksheet.cell(row=teststudent1_row, column=1).value, 'teststudent1@example.com')
            self.assertEqual(worksheet.cell(row=teststudent1_row, column=2).value, 1)
            self.assertEqual(worksheet.cell(row=teststudent1_row, column=3).value, 1)
            self.assertEqual(worksheet.cell(row=teststudent2_row, column=1).value, 'teststudent2@example.com')
            self.assertEqual(worksheet.cell(row=teststudent2_row, column=2).value, 1)
            self.assertEqual(worksheet.cell(row=teststudent2_row, column=3).value, 0)

    def test_grades_sheet_single_student_results_on_each_points_to_grade_mapper_type_assignment(self):
        requestuser = baker.make(settings.AUTH_USER_MODEL)
        period = self.create_single_student_results_on_each_points_to_grade_mapper_type_assignment()

        with mock.patch.object(DevilryReport, 'generator', all_results_generator.AllResultsExcelReportGenerator):
            worksheet = self.generate_report_and_get_worksheet(
                generated_by_user=requestuser, period=period, worksheet_name='Grades')
            self.assertEqual(worksheet.cell(row=2, column=1).value, 'teststudent@example.com')
            self.assertEqual(worksheet.cell(row=2, column=2).value, 'passed')
            self.assertEqual(worksheet.cell(row=2, column=3).value, '5/10')
            self.assertEqual(worksheet.cell(row=2, column=4).value, 'B')

    def test_points_sheet_single_student_results_on_each_points_to_grade_mapper_type_assignment(self):
        requestuser = baker.make(settings.AUTH_USER_MODEL)
        period = self.create_single_student_results_on_each_points_to_grade_mapper_type_assignment()

        with mock.patch.object(DevilryReport, 'generator', all_results_generator.AllResultsExcelReportGenerator):
            worksheet = self.generate_report_and_get_worksheet(
                generated_by_user=requestuser, period=period, worksheet_name='Points')
            self.assertEqual(worksheet.cell(row=2, column=1).value, 'teststudent@example.com')
            self.assertEqual(worksheet.cell(row=2, column=2).value, 5)
            self.assertEqual(worksheet.cell(row=2, column=3).value, 5)
            self.assertEqual(worksheet.cell(row=2, column=4).value, 47)

    def test_passed_sheet_single_student_results_on_each_points_to_grade_mapper_type_assignment(self):
        requestuser = baker.make(settings.AUTH_USER_MODEL)
        period = self.create_single_student_results_on_each_points_to_grade_mapper_type_assignment()

        with mock.patch.object(DevilryReport, 'generator', all_results_generator.AllResultsExcelReportGenerator):
            worksheet = self.generate_report_and_get_worksheet(
                generated_by_user=requestuser, period=period, worksheet_name='Passed Failed')
            self.assertEqual(worksheet.cell(row=2, column=1).value, 'teststudent@example.com')
            self.assertEqual(worksheet.cell(row=2, column=2).value, 1)
            self.assertEqual(worksheet.cell(row=2, column=3).value, 1)
            self.assertEqual(worksheet.cell(row=2, column=4).value, 1)
