# -*- coding: utf-8 -*-
from __future__ import unicode_literals

import mock
import openpyxl
from django import test
from django.conf import settings
from django.utils import timezone

from model_mommy import mommy
from io import BytesIO
from datetime import timedelta

from devilry.devilry_dbcache.customsql import AssignmentGroupDbCacheCustomSql
from devilry.devilry_group import devilry_group_mommy_factories as group_factory
from devilry.devilry_admin.views.period import all_results_generator
from devilry.devilry_report.models import DevilryReport


class AllResultsGeneratorPreMixin:
    def get_work_sheet(self, devilry_report):
        raise NotImplementedError()

    def make_assignment(self, period, **assignment_kwargs):
        return mommy.make('core.Assignment', parentnode=period, **assignment_kwargs)

    def make_relatedstudent(self, period, **relatedstudent_kwargs):
        return mommy.make('core.RelatedStudent', period=period, **relatedstudent_kwargs)

    def make_group_for_student(self, assignment, relatedstudent):
        group = mommy.make('core.AssignmentGroup', parentnode=assignment)
        mommy.make('core.Candidate', assignment_group=group, relatedstudent=relatedstudent)
        return group

    def test_headings(self):
        requestuser = mommy.make(settings.AUTH_USER_MODEL)
        period = mommy.make_recipe('devilry.apps.core.period_active')
        testassignment1 = self.make_assignment(period=period, long_name='Assignment 1')
        testassignment2 = self.make_assignment(period=period, long_name='Assignment 2')
        testassignment3 = self.make_assignment(period=period, long_name='Assignment 3')
        testassignment4 = self.make_assignment(period=period, long_name='Assignment 4')

        with mock.patch.object(DevilryReport, 'generator', all_results_generator.AllResultsExcelReportGenerator):
            devilry_report = DevilryReport(
                generator_options={'period_id': period.id},
                generator_type='semesterstudentresults',
                generated_by_user=requestuser)
            devilry_report.full_clean()
            devilry_report.save()
            devilry_report.generate()
            worksheet = self.get_work_sheet(devilry_report=devilry_report)
            self.assertEqual(worksheet.cell(row=0, column=0).value, 'Student')
            self.assertEqual(worksheet.cell(row=0, column=1).value, testassignment1.long_name)
            self.assertEqual(worksheet.cell(row=0, column=2).value, testassignment2.long_name)
            self.assertEqual(worksheet.cell(row=0, column=3).value, testassignment3.long_name)
            self.assertEqual(worksheet.cell(row=0, column=4).value, testassignment4.long_name)


class TestAllResultsGeneratorGradesSheet(AllResultsGeneratorPreMixin, test.TestCase):
    def setUp(self):
        AssignmentGroupDbCacheCustomSql().initialize()

    def get_work_sheet(self, devilry_report):
        return openpyxl.load_workbook(filename=BytesIO(devilry_report.result)).get_sheet_by_name(name='Grades')

    def test_value_for_single_student_passed_plugin_passed(self):
        requestuser = mommy.make(settings.AUTH_USER_MODEL)
        period = mommy.make_recipe('devilry.apps.core.period_active')
        testassignment = self.make_assignment(
            period=period, long_name='Assignment 1', passing_grade_min_points=1, max_points=1)
        teststudent = self.make_relatedstudent(period=period, user__shortname='teststudent@example.com')
        testgroup = self.make_group_for_student(assignment=testassignment, relatedstudent=teststudent)
        group_factory.feedbackset_first_attempt_published(group=testgroup, grading_points=1)

        with mock.patch.object(DevilryReport, 'generator', all_results_generator.AllResultsExcelReportGenerator):
            devilry_report = DevilryReport(
                generator_options={'period_id': period.id},
                generator_type='semesterstudentresults',
                generated_by_user=requestuser)
            devilry_report.full_clean()
            devilry_report.save()
            devilry_report.generate()
            worksheet = self.get_work_sheet(devilry_report=devilry_report)
            self.assertEqual(worksheet.cell(row=1, column=0).value, 'teststudent@example.com')
            self.assertEqual(worksheet.cell(row=1, column=1).value, 'passed')

    def test_value_for_single_student_failed_plugin_passed(self):
        requestuser = mommy.make(settings.AUTH_USER_MODEL)
        period = mommy.make_recipe('devilry.apps.core.period_active')
        testassignment = self.make_assignment(
            period=period, long_name='Assignment 1', passing_grade_min_points=1, max_points=1)
        teststudent = self.make_relatedstudent(period=period, user__shortname='teststudent@example.com')
        testgroup = self.make_group_for_student(assignment=testassignment, relatedstudent=teststudent)
        group_factory.feedbackset_first_attempt_published(group=testgroup, grading_points=0)

        with mock.patch.object(DevilryReport, 'generator', all_results_generator.AllResultsExcelReportGenerator):
            devilry_report = DevilryReport(
                generator_options={'period_id': period.id},
                generator_type='semesterstudentresults',
                generated_by_user=requestuser)
            devilry_report.full_clean()
            devilry_report.save()
            devilry_report.generate()
            worksheet = self.get_work_sheet(devilry_report=devilry_report)
            self.assertEqual(worksheet.cell(row=1, column=0).value, 'teststudent@example.com')
            self.assertEqual(worksheet.cell(row=1, column=1).value, 'failed')

    def test_value_for_single_student_plugin_raw_points(self):
        requestuser = mommy.make(settings.AUTH_USER_MODEL)
        period = mommy.make_recipe('devilry.apps.core.period_active')
        testassignment = self.make_assignment(
            period=period, long_name='Assignment 1', passing_grade_min_points=5, max_points=10,
            points_to_grade_mapper='raw-points')
        teststudent = self.make_relatedstudent(period=period, user__shortname='teststudent@example.com')
        testgroup = self.make_group_for_student(assignment=testassignment, relatedstudent=teststudent)
        group_factory.feedbackset_first_attempt_published(group=testgroup, grading_points=7)

        with mock.patch.object(DevilryReport, 'generator', all_results_generator.AllResultsExcelReportGenerator):
            devilry_report = DevilryReport(
                generator_options={'period_id': period.id},
                generator_type='semesterstudentresults',
                generated_by_user=requestuser)
            devilry_report.full_clean()
            devilry_report.save()
            devilry_report.generate()
            worksheet = self.get_work_sheet(devilry_report=devilry_report)
            self.assertEqual(worksheet.cell(row=1, column=0).value, 'teststudent@example.com')
            self.assertEqual(worksheet.cell(row=1, column=1).value, '7/10')

    def test_value_for_single_student_plugin_custom_table(self):
        requestuser = mommy.make(settings.AUTH_USER_MODEL)
        period = mommy.make_recipe('devilry.apps.core.period_active')
        testassignment = self.make_assignment(
            period=period, long_name='Assignment 1', passing_grade_min_points=5, max_points=10,
            points_to_grade_mapper='custom-table')
        point_to_grade_map = mommy.make('core.PointToGradeMap', assignment=testassignment)
        mommy.make('core.PointRangeToGrade',point_to_grade_map=point_to_grade_map,
                   minimum_points=0, maximum_points=10, grade='F')
        mommy.make('core.PointRangeToGrade', point_to_grade_map=point_to_grade_map,
                   minimum_points=11, maximum_points=20, grade='E')
        mommy.make('core.PointRangeToGrade', point_to_grade_map=point_to_grade_map,
                   minimum_points=21, maximum_points=30, grade='D')
        mommy.make('core.PointRangeToGrade', point_to_grade_map=point_to_grade_map,
                   minimum_points=31, maximum_points=40, grade='C')
        mommy.make('core.PointRangeToGrade', point_to_grade_map=point_to_grade_map,
                   minimum_points=41, maximum_points=50, grade='B')
        mommy.make('core.PointRangeToGrade', point_to_grade_map=point_to_grade_map,
                   minimum_points=51, maximum_points=60, grade='A')

        teststudent = self.make_relatedstudent(period=period, user__shortname='teststudent@example.com')
        testgroup = self.make_group_for_student(assignment=testassignment, relatedstudent=teststudent)
        group_factory.feedbackset_first_attempt_published(group=testgroup, grading_points=32)

        with mock.patch.object(DevilryReport, 'generator', all_results_generator.AllResultsExcelReportGenerator):
            devilry_report = DevilryReport(
                generator_options={'period_id': period.id},
                generator_type='semesterstudentresults',
                generated_by_user=requestuser)
            devilry_report.full_clean()
            devilry_report.save()
            devilry_report.generate()
            worksheet = self.get_work_sheet(devilry_report=devilry_report)
            self.assertEqual(worksheet.cell(row=1, column=0).value, 'teststudent@example.com')
            self.assertEqual(worksheet.cell(row=1, column=1).value, 'C')

    def test_value_for_single_student_not_registered_on_assignment(self):
        requestuser = mommy.make(settings.AUTH_USER_MODEL)
        period = mommy.make_recipe('devilry.apps.core.period_active')
        self.make_assignment(period=period, long_name='Assignment 1', passing_grade_min_points=1, max_points=1)
        self.make_relatedstudent(period=period, user__shortname='teststudent@example.com')

        with mock.patch.object(DevilryReport, 'generator', all_results_generator.AllResultsExcelReportGenerator):
            devilry_report = DevilryReport(
                generator_options={'period_id': period.id},
                generator_type='semesterstudentresults',
                generated_by_user=requestuser)
            devilry_report.full_clean()
            devilry_report.save()
            devilry_report.generate()
            worksheet = self.get_work_sheet(devilry_report=devilry_report)
            self.assertEqual(worksheet.cell(row=1, column=0).value, 'teststudent@example.com')
            self.assertEqual(worksheet.cell(row=1, column=1).value, 'Not registered')

    def test_value_for_single_student_waiting_for_feedback(self):
        requestuser = mommy.make(settings.AUTH_USER_MODEL)
        period = mommy.make_recipe('devilry.apps.core.period_active')
        testassignment = self.make_assignment(period=period, long_name='Assignment 1', passing_grade_min_points=1, max_points=1)
        teststudent = self.make_relatedstudent(period=period, user__shortname='teststudent@example.com')
        testgroup = self.make_group_for_student(assignment=testassignment, relatedstudent=teststudent)
        deadline_datetime = timezone.now() - timezone.timedelta(days=10)
        group_factory.feedbackset_first_attempt_unpublished(group=testgroup, deadline_datetime=deadline_datetime)

        with mock.patch.object(DevilryReport, 'generator', all_results_generator.AllResultsExcelReportGenerator):
            devilry_report = DevilryReport(
                generator_options={'period_id': period.id},
                generator_type='semesterstudentresults',
                generated_by_user=requestuser)
            devilry_report.full_clean()
            devilry_report.save()
            devilry_report.generate()
            worksheet = self.get_work_sheet(devilry_report=devilry_report)
            self.assertEqual(worksheet.cell(row=1, column=0).value, 'teststudent@example.com')
            self.assertEqual(worksheet.cell(row=1, column=1).value, 'Waiting for feedback')

    def test_value_for_single_student_hard_deadline_no_deliveries(self):
        requestuser = mommy.make(settings.AUTH_USER_MODEL)
        period = mommy.make_recipe('devilry.apps.core.period_active')
        testassignment = self.make_assignment(
            period=period, long_name='Assignment 1', passing_grade_min_points=1, max_points=1, deadline_handling=1)
        teststudent = self.make_relatedstudent(period=period, user__shortname='teststudent@example.com')
        testgroup = self.make_group_for_student(assignment=testassignment, relatedstudent=teststudent)
        deadline_datetime = timezone.now() - timezone.timedelta(days=10)
        group_factory.feedbackset_first_attempt_unpublished(group=testgroup, deadline_datetime=deadline_datetime)

        with mock.patch.object(DevilryReport, 'generator', all_results_generator.AllResultsExcelReportGenerator):
            devilry_report = DevilryReport(
                generator_options={'period_id': period.id},
                generator_type='semesterstudentresults',
                generated_by_user=requestuser)
            devilry_report.full_clean()
            devilry_report.save()
            devilry_report.generate()
            worksheet = self.get_work_sheet(devilry_report=devilry_report)
            self.assertEqual(worksheet.cell(row=1, column=0).value, 'teststudent@example.com')
            self.assertEqual(worksheet.cell(row=1, column=1).value, 'No deliveries')

    def test_value_for_single_student_hard_deadline_waiting_for_feedback(self):
        requestuser = mommy.make(settings.AUTH_USER_MODEL)
        period = mommy.make_recipe('devilry.apps.core.period_active')
        testassignment = self.make_assignment(
            period=period, long_name='Assignment 1', passing_grade_min_points=1, max_points=1, deadline_handling=1)
        teststudent = self.make_relatedstudent(period=period, user__shortname='teststudent@example.com')
        testgroup = self.make_group_for_student(assignment=testassignment, relatedstudent=teststudent)
        deadline_datetime = timezone.now() - timezone.timedelta(days=10)
        feedbackset = group_factory.feedbackset_first_attempt_unpublished(group=testgroup, deadline_datetime=deadline_datetime)
        mommy.make('devilry_group.GroupComment', feedback_set=feedbackset, user=teststudent.user,
                   user_role='student', text='Test')

        with mock.patch.object(DevilryReport, 'generator', all_results_generator.AllResultsExcelReportGenerator):
            devilry_report = DevilryReport(
                generator_options={'period_id': period.id},
                generator_type='semesterstudentresults',
                generated_by_user=requestuser)
            devilry_report.full_clean()
            devilry_report.save()
            devilry_report.generate()
            worksheet = self.get_work_sheet(devilry_report=devilry_report)
            self.assertEqual(worksheet.cell(row=1, column=0).value, 'teststudent@example.com')
            self.assertEqual(worksheet.cell(row=1, column=1).value, 'Waiting for feedback')

    def test_value_for_single_student_waiting_for_deliveries(self):
        requestuser = mommy.make(settings.AUTH_USER_MODEL)
        period = mommy.make_recipe('devilry.apps.core.period_active')
        testassignment = self.make_assignment(
            period=period, long_name='Assignment 1', passing_grade_min_points=1, max_points=1, deadline_handling=1)
        teststudent = self.make_relatedstudent(period=period, user__shortname='teststudent@example.com')
        testgroup = self.make_group_for_student(assignment=testassignment, relatedstudent=teststudent)
        deadline_datetime = timezone.now() + timezone.timedelta(days=10)
        group_factory.feedbackset_first_attempt_unpublished(group=testgroup, deadline_datetime=deadline_datetime)

        with mock.patch.object(DevilryReport, 'generator', all_results_generator.AllResultsExcelReportGenerator):
            devilry_report = DevilryReport(
                generator_options={'period_id': period.id},
                generator_type='semesterstudentresults',
                generated_by_user=requestuser)
            devilry_report.full_clean()
            devilry_report.save()
            devilry_report.generate()
            worksheet = self.get_work_sheet(devilry_report=devilry_report)
            self.assertEqual(worksheet.cell(row=1, column=0).value, 'teststudent@example.com')
            self.assertEqual(worksheet.cell(row=1, column=1).value, 'Waiting for deliveries')

    def test_value_for_single_student_on_multiple_assignments(self):
        requestuser = mommy.make(settings.AUTH_USER_MODEL)
        period = mommy.make_recipe('devilry.apps.core.period_active')
        testassignment1 = self.make_assignment(
            period=period, long_name='Assignment 1', passing_grade_min_points=1, max_points=1)
        testassignment2 = self.make_assignment(
            period=period, long_name='Assignment 1', passing_grade_min_points=1, max_points=1)
        testassignment3 = self.make_assignment(
            period=period, long_name='Assignment 1', passing_grade_min_points=1, max_points=1)
        teststudent = self.make_relatedstudent(period=period, user__shortname='teststudent@example.com')
        testgroup1 = self.make_group_for_student(assignment=testassignment1, relatedstudent=teststudent)
        testgroup2 = self.make_group_for_student(assignment=testassignment2, relatedstudent=teststudent)
        testgroup3 = self.make_group_for_student(assignment=testassignment3, relatedstudent=teststudent)
        group_factory.feedbackset_first_attempt_published(group=testgroup1, grading_points=1)
        group_factory.feedbackset_first_attempt_published(group=testgroup2, grading_points=1)
        group_factory.feedbackset_first_attempt_published(group=testgroup3, grading_points=0)

        with mock.patch.object(DevilryReport, 'generator', all_results_generator.AllResultsExcelReportGenerator):
            devilry_report = DevilryReport(
                generator_options={'period_id': period.id},
                generator_type='semesterstudentresults',
                generated_by_user=requestuser)
            devilry_report.full_clean()
            devilry_report.save()
            devilry_report.generate()
            worksheet = self.get_work_sheet(devilry_report=devilry_report)
            self.assertEqual(worksheet.cell(row=1, column=0).value, 'teststudent@example.com')
            self.assertEqual(worksheet.cell(row=1, column=1).value, 'passed')
            self.assertEqual(worksheet.cell(row=1, column=2).value, 'passed')
            self.assertEqual(worksheet.cell(row=1, column=3).value, 'failed')

    def test_multiple_students_on_multiple_assignments(self):
        requestuser = mommy.make(settings.AUTH_USER_MODEL)
        period = mommy.make_recipe('devilry.apps.core.period_active')
        testassignment1 = self.make_assignment(
            period=period, long_name='Assignment 1', passing_grade_min_points=1, max_points=1)
        testassignment2 = self.make_assignment(
            period=period, long_name='Assignment 1', passing_grade_min_points=1, max_points=1)
        teststudent1 = self.make_relatedstudent(period=period, user__shortname='teststudent1@example.com')
        teststudent2 = self.make_relatedstudent(period=period, user__shortname='teststudent2@example.com')
        testgroup1 = self.make_group_for_student(assignment=testassignment1, relatedstudent=teststudent1)
        testgroup2 = self.make_group_for_student(assignment=testassignment2, relatedstudent=teststudent1)
        testgroup3 = self.make_group_for_student(assignment=testassignment1, relatedstudent=teststudent2)
        testgroup4 = self.make_group_for_student(assignment=testassignment2, relatedstudent=teststudent2)
        group_factory.feedbackset_first_attempt_published(group=testgroup1, grading_points=1)
        group_factory.feedbackset_first_attempt_published(group=testgroup2, grading_points=1)
        group_factory.feedbackset_first_attempt_published(group=testgroup3, grading_points=1)
        group_factory.feedbackset_first_attempt_published(group=testgroup4, grading_points=0)

        with mock.patch.object(DevilryReport, 'generator', all_results_generator.AllResultsExcelReportGenerator):
            devilry_report = DevilryReport(
                generator_options={'period_id': period.id},
                generator_type='semesterstudentresults',
                generated_by_user=requestuser)
            devilry_report.full_clean()
            devilry_report.save()
            devilry_report.generate()
            worksheet = self.get_work_sheet(devilry_report=devilry_report)
            self.assertEqual(worksheet.cell(row=1, column=0).value, 'teststudent1@example.com')
            self.assertEqual(worksheet.cell(row=1, column=1).value, 'passed')
            self.assertEqual(worksheet.cell(row=1, column=2).value, 'passed')
            self.assertEqual(worksheet.cell(row=2, column=0).value, 'teststudent2@example.com')
            self.assertEqual(worksheet.cell(row=2, column=1).value, 'passed')
            self.assertEqual(worksheet.cell(row=2, column=2).value, 'failed')
